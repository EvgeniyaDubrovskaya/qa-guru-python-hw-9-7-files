import os
from io import BytesIO, TextIOWrapper
from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook
CUR_DIR = os.path.dirname(os.path.abspath(__file__))


def test_check_file_context():
    # Pack files
    tmp_dir = os.path.join(CUR_DIR, "tmp")
    res_dir = os.path.join(CUR_DIR, "resources")
    with ZipFile(res_dir + "\\files.zip", mode="w") as archive:
        for file in os.listdir(tmp_dir):
            archive.write(os.path.join(tmp_dir, file), str.lower(file.title()))

    # Open archive
    zip = ZipFile(res_dir + "\\files.zip", mode="r")

    # Read pdf files from archive
    pdf_reader = PdfReader(BytesIO(zip.read("sample_pdf.pdf")))
    assert "A Simple PDF File" in pdf_reader.pages[0].extract_text()

    # Read csv files from archive
    csv_readr = list(csv.reader(TextIOWrapper(zip.open("sample_csv.csv"), 'utf-8')))

    assert "John" in csv_readr[0][0]
    assert 6 == len(csv_readr)

    # Read xlsx files from archive
    workbook = load_workbook(BytesIO(zip.read("sample_xlsx.xlsx")))
    sheet = workbook.active
    assert "Dulce" in sheet.cell(row=2, column=2).value
